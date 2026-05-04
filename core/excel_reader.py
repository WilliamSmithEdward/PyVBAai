# Copyright (c) 2026 William E. Smith (williamsmithe@icloud.com). All rights reserved.
# Proprietary and confidential. Unauthorized use, copying, or distribution
# of this file, via any medium, is strictly prohibited. See LICENSE for details.
"""Excel workbook extraction.

Cell values, formulas and formatting are read directly from the xlsx/xlsm XML via
openpyxl -- no Excel process required.

VBA source code cannot be read from the binary OLE blob without COM, so that path
still uses win32com but is only invoked for .xlsm/.xlam files.
"""
from __future__ import annotations

import os
from functools import lru_cache
from typing import Any

from models.workbook import (
    CellData,
    CellFormat,
    ContextConfig,
    NamedRange,
    SheetData,
    VBAModule,
    WorkbookData,
)

# Maximum columns we ever read per sheet (sanity limit)
_MAX_COLS_HARD = 200


@lru_cache(maxsize=512)
def col_letter(col: int) -> str:
    """1-based column index -> letter(s), e.g. 1->'A', 27->'AA'. Cached."""
    result = ""
    while col > 0:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result


def cell_address(row: int, col: int) -> str:
    return f"{col_letter(col)}{row}"


def _abs_addr(row: int, col: int) -> str:
    return f"${col_letter(col)}${row}"


# -- connected-component flood-fill --------------------------------------------

def _connected_area_addresses(cells: dict) -> list[str]:
    """Return Excel-style absolute addresses for each connected block of non-empty cells."""
    occupied: set[tuple[int, int]] = {(cd.row, cd.col) for cd in cells.values()}
    visited: set[tuple[int, int]] = set()
    areas: list[str] = []

    for seed in sorted(occupied):
        if seed in visited:
            continue
        component: list[tuple[int, int]] = []
        queue = [seed]
        visited.add(seed)
        while queue:
            r, c = queue.pop()
            component.append((r, c))
            for nr, nc in ((r - 1, c), (r + 1, c), (r, c - 1), (r, c + 1)):
                if (nr, nc) in occupied and (nr, nc) not in visited:
                    visited.add((nr, nc))
                    queue.append((nr, nc))
        min_r = min(p[0] for p in component)
        max_r = max(p[0] for p in component)
        min_c = min(p[1] for p in component)
        max_c = max(p[1] for p in component)
        areas.append(f"{_abs_addr(min_r, min_c)}:{_abs_addr(max_r, max_c)}")

    return areas


# -- format helpers ------------------------------------------------------------

def _argb_to_hex(argb: str | None) -> str:
    """Convert openpyxl ARGB string 'FF112233' -> '112233'. Returns '' if default."""
    if not argb or argb in ("00000000", "FFFFFFFF", "FF000000"):
        return ""
    return argb[2:] if len(argb) == 8 else argb


def _cell_format(cell: Any) -> CellFormat | None:
    """Extract a CellFormat from an openpyxl cell. Returns None if all defaults."""
    fmt = CellFormat()
    changed = False

    nf = getattr(cell, "number_format", None)
    if nf and nf != "General":
        fmt.number_format = nf
        changed = True

    font = getattr(cell, "font", None)
    if font:
        if font.bold:
            fmt.bold = True
            changed = True
        if font.italic:
            fmt.italic = True
            changed = True
        if font.underline and font.underline != "none":
            fmt.underline = True
            changed = True
        if font.name and font.name != "Calibri":
            fmt.font_name = font.name
            changed = True
        if font.size and font.size != 11:
            fmt.font_size = float(font.size)
            changed = True
        try:
            clr = font.color
            if clr and clr.type == "rgb":
                h = _argb_to_hex(clr.rgb)
                if h:
                    fmt.font_color = h
                    changed = True
        except Exception:  # noqa: BLE001
            pass

    fill = getattr(cell, "fill", None)
    if fill and fill.fill_type not in (None, "none"):
        try:
            fg = fill.fgColor
            if fg and fg.type == "rgb":
                h = _argb_to_hex(fg.rgb)
                if h:
                    fmt.bg_color = h
                    changed = True
        except Exception:  # noqa: BLE001
            pass

    alignment = getattr(cell, "alignment", None)
    if alignment:
        if alignment.horizontal and alignment.horizontal != "general":
            fmt.h_align = alignment.horizontal
            changed = True
        if alignment.vertical and alignment.vertical != "bottom":
            fmt.v_align = alignment.vertical
            changed = True
        if alignment.wrap_text:
            fmt.wrap_text = True
            changed = True

    protection = getattr(cell, "protection", None)
    if protection and not protection.locked:
        fmt.locked = False
        changed = True

    return fmt if changed else None


# -- main entry point ----------------------------------------------------------

def read_workbook(file_path: str, config: ContextConfig | None = None) -> WorkbookData:
    """
    Extract workbook data using openpyxl for cell data and formatting.
    VBA source code is read via COM only for .xlsm/.xlam files.
    No Excel process is launched for plain .xlsx files.
    """
    import openpyxl

    if config is None:
        config = ContextConfig()

    name = os.path.basename(file_path)
    wb_data = WorkbookData(file_path=file_path, name=name)
    try:
        wb_data.loaded_mtime = os.path.getmtime(file_path)
    except OSError:
        pass

    try:
        owb = openpyxl.load_workbook(
            file_path,
            read_only=False,
            data_only=False,
            keep_vba=True,
            keep_links=False,
        )

        # -- Sheets -----------------------------------------------------------
        for idx, ws in enumerate(owb.worksheets):
            visible = ws.sheet_state == "visible"

            if (
                config.included_sheets is not None
                and ws.title not in config.included_sheets
            ):
                wb_data.sheets.append(SheetData(
                    index=idx, name=ws.title,
                    used_range_address="", row_count=0, col_count=0,
                    visible=visible,
                ))
                continue

            min_row = ws.min_row or 1
            min_col = ws.min_column or 1
            max_row = ws.max_row or 0
            max_col = min(ws.max_column or 0, _MAX_COLS_HARD)
            row_count = max(0, max_row - min_row + 1) if max_row else 0
            col_count = max(0, max_col - min_col + 1) if max_col else 0

            if max_row and max_col:
                used_addr = (
                    f"${col_letter(min_col)}${min_row}"
                    f":${col_letter(max_col)}${max_row}"
                )
            else:
                used_addr = ""

            sheet_data = SheetData(
                index=idx,
                name=ws.title,
                used_range_address=used_addr,
                row_count=row_count,
                col_count=col_count,
                visible=visible,
            )

            if max_row and max_col:
                for row in ws.iter_rows(
                    min_row=min_row, max_row=max_row,
                    min_col=min_col, max_col=max_col,
                ):
                    for cell in row:
                        if cell.value is None:
                            continue
                        r, c = cell.row, cell.column
                        addr = cell_address(r, c)
                        formula = ""
                        value = cell.value
                        if isinstance(value, str) and value.startswith("="):
                            formula = value
                            value = None
                        sheet_data.cells[addr] = CellData(
                            row=r, col=c, address=addr,
                            value=value, formula=formula,
                            fmt=_cell_format(cell),
                        )

            wb_data.sheets.append(sheet_data)
            sheet_data.area_addresses = _connected_area_addresses(sheet_data.cells)

        # -- Named ranges -----------------------------------------------------
        if config.include_named_ranges:
            for defn in owb.defined_names:
                try:
                    wb_data.named_ranges.append(
                        NamedRange(name=defn.name, refers_to=defn.value)
                    )
                except Exception:  # noqa: BLE001
                    pass

        owb.close()

        # -- VBA source code -- COM path (xlsm/xlam only) ---------------------
        ext = os.path.splitext(file_path)[1].lower()
        if config.include_vba and ext in (".xlsm", ".xlam"):
            _read_vba_via_com(file_path, wb_data, config)

    except Exception as exc:
        wb_data.extraction_error = str(exc)

    return wb_data


def _read_vba_via_com(
    file_path: str, wb_data: WorkbookData, config: ContextConfig
) -> None:
    """Populate wb_data.vba_modules using COM. Called only for .xlsm/.xlam files."""
    try:
        import pythoncom
        import win32com.client as win32
    except ImportError:
        wb_data.extraction_error = "pywin32 is required to read VBA from .xlsm files."
        return

    pythoncom.CoInitialize()
    excel = None
    com_wb = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = True   # must be visible during Open -- headless blocks on OneDrive sync
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False
        excel.AutomationSecurity = 3  # msoAutomationSecurityForceDisable
        try:
            excel.Calculation = -4135
        except Exception:  # noqa: BLE001
            pass

        com_wb = excel.Workbooks.Open(
            file_path, UpdateLinks=False, ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
        )
        excel.Visible = False

        type_map = {1: "Module", 2: "Class", 3: "Form", 100: "Document"}
        for comp in com_wb.VBProject.VBComponents:
            try:
                mname = comp.Name
                if (
                    config.included_vba_modules is not None
                    and mname not in config.included_vba_modules
                ):
                    continue
                if mname in config.excluded_vba_modules:
                    continue
                lines = comp.CodeModule.CountOfLines
                code = comp.CodeModule.Lines(1, lines) if lines > 0 else ""
                wb_data.vba_modules.append(VBAModule(
                    name=mname,
                    module_type=comp.Type,
                    type_name=type_map.get(comp.Type, "Unknown"),
                    code=code,
                ))
                wb_data.has_vba = True
            except Exception:  # noqa: BLE001
                pass

    except Exception as exc:
        wb_data.extraction_error = (
            f"VBA extraction failed: {exc}. "
            "Enable 'Trust access to the VBA project object model' in "
            "Excel > Trust Center > Macro Settings."
        )
    finally:
        try:
            if com_wb is not None:
                com_wb.Close(SaveChanges=False)
        except Exception:  # noqa: BLE001
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:  # noqa: BLE001
            pass
        pythoncom.CoUninitialize()
