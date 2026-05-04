# Copyright (c) 2026 William E. Smith (williamsmithe@icloud.com). All rights reserved.
# Proprietary and confidential. Unauthorized use, copying, or distribution
# of this file, via any medium, is strictly prohibited. See LICENSE for details.
"""Apply AI-generated changes to an Excel workbook.

All operations use openpyxl directly -- no Excel process needed.
VBA operations (set_vba, add_vba_module, delete_vba_module) are handled
by the UI as copy/paste and are never passed to this module.
"""
from __future__ import annotations

import os
import re
import shutil
import tempfile
import zipfile
from typing import Any

from app.logger import get_logger
from models.conversation import Change

_log = get_logger(__name__)

# Functions added after Excel 2010 that require an _xlfn. prefix in OOXML.
# Without this prefix openpyxl-saved files show #NAME? when opened in Excel.
_XLFN_FUNCTIONS: frozenset[str] = frozenset({
    # Excel 365 dynamic array / spill
    "FILTER", "UNIQUE", "SORT", "SORTBY", "SEQUENCE", "RANDARRAY",
    # Excel 365 lookup
    "XLOOKUP", "XMATCH",
    # Excel 365 lambda / let
    "LET", "LAMBDA", "MAP", "REDUCE", "SCAN", "MAKEARRAY", "BYROW", "BYCOL", "ISOMITTED",
    # Excel 2019
    "IFS", "SWITCH", "MAXIFS", "MINIFS", "CONCAT", "TEXTJOIN", "STOCKHISTORY",
    # Excel 2013
    "ACOT", "ACOTH", "ARABIC", "BASE", "BITAND", "BITLSHIFT", "BITOR", "BITRSHIFT",
    "BITXOR", "CEILING.MATH", "COMBINA", "COT", "COTH", "CSC", "CSCH", "DAYS",
    "DECIMAL", "ENCODEURL", "FILTERXML", "FLOOR.MATH", "FORMULATEXT", "GAMMA",
    "GAUSS", "IFNA", "IMCOSH", "IMCOT", "IMCSC", "IMCSCH", "IMSEC", "IMSECH",
    "IMSINH", "IMTAN", "ISFORMULA", "MUNIT", "NUMBERVALUE", "PDURATION",
    "PERMUTATIONA", "PHI", "RRI", "SEC", "SECH", "SHEET", "SHEETS", "SKEW.P",
    "UNICHAR", "UNICODE", "WEBSERVICE", "XOR",
    # Excel 2010
    "AGGREGATE", "NETWORKDAYS.INTL", "WORKDAY.INTL",
})

_XLFN_RE = re.compile(
    r"(?<![._A-Za-z0-9])("
    + "|".join(re.escape(fn) for fn in sorted(_XLFN_FUNCTIONS, key=len, reverse=True))
    + r")(?=\s*\()",
    re.IGNORECASE,
)

# Spilled-range operator: "A1#" in the formula bar must be stored as
# _xlfn.ANCHORARRAY(A1) in OOXML. Storing the literal "#" form crashes Excel
# with "Removed Records: Formula" on next open. Match an optional sheet
# qualifier ("Sheet1!" or "'My Sheet'!") followed by an A1-style address and
# a trailing "#". Avoid matching inside larger identifiers via lookbehind.
_SPILL_REF_RE = re.compile(
    r"(?<![A-Za-z0-9_.])"
    r"((?:'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?"   # optional sheet qualifier
    r"(\$?[A-Z]+\$?\d+)#",
)


def _add_xlfn_prefix(formula: str) -> str:
    """Rewrite formula so Excel-2013+ functions carry the required _xlfn. prefix.

    openpyxl stores formulas verbatim, but the OOXML spec requires that
    functions introduced after Excel 2010 be stored as _xlfn.FUNCNAME().
    Without the prefix Excel shows #NAME? and prepends @ to suppress spilling.
    Already-prefixed functions are left unchanged (lookbehind skips after '.').

    Also rewrites the spilled-range operator "A1#" to its OOXML form
    ``_xlfn.ANCHORARRAY(A1)``; storing the bare "#" makes Excel discard the
    formula on next open.
    """
    if not formula.startswith("="):
        return formula
    body = formula[1:]
    body = _XLFN_RE.sub(lambda m: f"_xlfn.{m.group(1).upper()}", body)
    body = _SPILL_REF_RE.sub(
        lambda m: f"_xlfn.ANCHORARRAY({(m.group(1) or '')}{m.group(2)})",
        body,
    )
    return "=" + body


class ApplyError(Exception):
    """Raised when a change operation fails."""


# xl/workbook.xml extension that tells Excel 365 this workbook understands
# dynamic arrays.  Without it Excel adds "@" (implicit intersection operator)
# to every formula that can spill, because it assumes the file was saved by a
# pre-365 application that never intended spilling behaviour.
_DA_FEATURES_URI = "{B58B0392-4F1F-4190-BB64-5DF3571DCE5F}"
_DA_FEATURES_EXT = (
    f'<ext uri="{_DA_FEATURES_URI}"'
    ' xmlns:xcalcf="http://schemas.microsoft.com/office/spreadsheetml/2018/calcfeatures">'
    "<xcalcf:calcFeatures>"
    '<xcalcf:feature name="microsoft.com:RD"/>'
    '<xcalcf:feature name="microsoft.com:Single"/>'
    '<xcalcf:feature name="microsoft.com:FV"/>'
    '<xcalcf:feature name="microsoft.com:CNMTM"/>'
    '<xcalcf:feature name="microsoft.com:LET_WF"/>'
    '<xcalcf:feature name="microsoft.com:LAMBDA_WF"/>'
    '<xcalcf:feature name="microsoft.com:ARRAYTEXT_WF"/>'
    "</xcalcf:calcFeatures>"
    "</ext>"
)

# xl/metadata.xml that registers the XLDAPR dynamic-array cell-metadata type.
# Every anchor cell for a spilling formula must reference this via cm="1".
# Excel requires this file to exist (+ content-type + relationship) before it
# will treat formulas as dynamic arrays instead of implicit-intersection (@).
_DA_METADATA_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<metadata xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
    ' xmlns:xda="http://schemas.microsoft.com/office/spreadsheetml/2017/dynamicarray">'
    '<metadataTypes count="1">'
    '<metadataType name="XLDAPR" minSupportedVersion="120000" copy="1" pasteAll="1"'
    ' pasteValues="1" merge="1" splitFirst="1" rowColShift="1" clearFormats="1"'
    ' clearComments="1" assign="1" coerce="1" cellMeta="1"/>'
    "</metadataTypes>"
    '<futureMetadata name="XLDAPR" count="1">'
    "<bk><extLst>"
    '<ext uri="{bdbb8cdc-fa1e-496e-a857-3c3f30c029c3}">'
    '<xda:dynamicArrayProperties fDynamic="1" fCollapsed="0"/>'
    "</ext></extLst></bk>"
    "</futureMetadata>"
    '<cellMetadata count="1"><bk><rc t="1" v="0"/></bk></cellMetadata>'
    "</metadata>"
)
_DA_CONTENT_TYPE = (
    '<Override PartName="/xl/metadata.xml"'
    ' ContentType="application/vnd.openxmlformats-officedocument'
    '.spreadsheetml.sheetMetadata+xml"/>'
)
_DA_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sheetMetadata"
)

# Dynamic-array function names that need the cm="1" + t="array" treatment.
_DA_FUNC_RE = (
    r"_xlfn\.(?:FILTER|UNIQUE|SORT(?:BY)?|SEQUENCE|RANDARRAY"
    r"|XLOOKUP|XMATCH|LET|LAMBDA|MAP|REDUCE|SCAN|MAKEARRAY|BYROW|BYCOL)"
)

# Pass 1 - new dynamic-array formulas written by openpyxl as plain <f>:
# <c r="A1"><f>_xlfn.FILTER(...)</f> -> stamp cm="1" + t="array" ref="A1".
# We use the single anchor cell as ref and rely on fullCalcOnLoad="1" to make
# Excel resize the spill range on open.
_DA_NEW_RE = re.compile(
    r'(<c\b(?![^>]*\bcm=)[^>]*\br="([A-Z]+\d+)"[^>]*>)'
    r"(<f\b(?![^>]*\bt=)[^>]*>)"
    r"(" + _DA_FUNC_RE + r")",
    re.IGNORECASE,
)

# Pass 2 - existing dynamic-array formulas re-saved by openpyxl: openpyxl
# preserves <f t="array" ref="..."> but drops cm="1" on the <c> element AND
# drops xl/metadata.xml. The postprocessor restores the metadata part; this
# regex re-adds cm="1" to the anchor cell so Excel keeps treating it as a
# dynamic array instead of converting it back to a legacy CSE array.
_DA_EXISTING_RE = re.compile(
    r'(<c\b(?![^>]*\bcm=)[^>]*>)'
    r'(<f\b[^>]*\bt="array"[^>]*>)'
    r"(" + _DA_FUNC_RE + r")",
    re.IGNORECASE,
)


def _da_new_sub(m: re.Match) -> str:  # type: ignore[type-arg]
    c_open = m.group(1)[:-1]   # strip trailing >
    anchor = m.group(2)
    f_open = m.group(3)[:-1]   # strip trailing >
    return f'{c_open} cm="1">{f_open} t="array" ref="{anchor}">{m.group(4)}'


def _da_existing_sub(m: re.Match) -> str:  # type: ignore[type-arg]
    c_open = m.group(1)[:-1]   # strip trailing >
    return f'{c_open} cm="1">{m.group(2)}{m.group(3)}'


def _postprocess_xlsx(xlsx_path: str) -> None:
    """Single-pass zip fixup applied after every openpyxl save.

    1. Remove xl/calcChain.xml -- stale chain causes "Removed Records: Formula"
       repair dialogs; Excel regenerates it on next open.
    2. Inject xcalcf:calcFeatures into xl/workbook.xml so Excel 365 marks the
       workbook as dynamic-array aware.
    3. Add xl/metadata.xml (XLDAPR) + its content-type + relationship so Excel
       recognises per-cell dynamic-array intent.
    4. Add cm="1" / t="array" to anchor <c>/<f> elements for spilling functions
       so Excel does not prepend the @ implicit-intersection operator.
    """
    tmp = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as zin, \
         zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        names = set(zin.namelist())
        need_metadata = "xl/metadata.xml" not in names

        for name in names:
            if name == "xl/calcChain.xml":
                continue
            data = zin.read(name)
            text = data.decode("utf-8")

            if name == "xl/workbook.xml":
                if _DA_FEATURES_URI not in text:
                    if "</extLst>" in text:
                        text = text.replace("</extLst>", _DA_FEATURES_EXT + "</extLst>", 1)
                    else:
                        text = text.replace(
                            "</workbook>",
                            "<extLst>" + _DA_FEATURES_EXT + "</extLst></workbook>",
                            1,
                        )
                data = text.encode("utf-8")

            elif name == "[Content_Types].xml" and need_metadata:
                if "/xl/metadata.xml" not in text:
                    text = text.replace("</Types>", _DA_CONTENT_TYPE + "</Types>", 1)
                data = text.encode("utf-8")

            elif name == "xl/_rels/workbook.xml.rels" and need_metadata:
                if _DA_REL_TYPE not in text:
                    existing = [int(n) for n in re.findall(r'Id="rId(\d+)"', text)]
                    next_id = max(existing, default=0) + 1
                    new_rel = (
                        f'<Relationship Id="rId{next_id}" Type="{_DA_REL_TYPE}"'
                        ' Target="metadata.xml"/>'
                    )
                    text = text.replace("</Relationships>", new_rel + "</Relationships>", 1)
                data = text.encode("utf-8")

            elif name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                text = _DA_NEW_RE.sub(_da_new_sub, text)
                text = _DA_EXISTING_RE.sub(_da_existing_sub, text)
                data = text.encode("utf-8")

            zout.writestr(name, data)

        if need_metadata:
            zout.writestr("xl/metadata.xml", _DA_METADATA_XML)

    os.replace(tmp, xlsx_path)


def apply_changes(file_path: str, changes: list[Change]) -> str:
    """
    Apply every change to *file_path* and save.

    Returns the actual saved path (may differ from file_path if .xlsx was
    promoted to .xlsm due to VBA changes).

    Raises ApplyError on the first failure.
    Callers should create a backup BEFORE calling this function.
    """
    # Fail fast with a clear message if the file is locked.
    try:
        with open(file_path, "r+b"):
            pass
    except PermissionError as exc:
        fname = os.path.basename(file_path)
        raise ApplyError(
            f"'{fname}' is open in another program. Close it and try again."
        ) from exc
    except OSError as exc:
        raise ApplyError(f"Cannot access '{os.path.basename(file_path)}': {exc}") from exc

    # ── Apply all changes via openpyxl ────────────────────────────────────────
    import openpyxl
    saved_path = os.path.abspath(file_path)
    # Only preserve VBA for macro-enabled workbooks (.xlsm/.xlam).
    # Using keep_vba=True on a plain .xlsx causes openpyxl to embed a
    # vbaProject.bin and write xlsm-style content types, which makes Excel
    # reject the file as having an invalid format/extension mismatch.
    is_macro_enabled = os.path.splitext(file_path)[1].lower() in (".xlsm", ".xlam")
    owb = openpyxl.load_workbook(file_path, keep_vba=is_macro_enabled, keep_links=False)
    try:
        for change in changes:
            _dispatch_openpyxl(owb, change)

        # Force full recalculation on open so dynamic array formulas
        # (FILTER, XLOOKUP, UNIQUE, SORT, etc.) spill without requiring F2+Enter.
        owb.calculation.fullCalcOnLoad = True

        tmp_dir = tempfile.mkdtemp()
        try:
            tmp_path = os.path.join(tmp_dir, os.path.basename(saved_path))
            owb.save(tmp_path)
            # Remove stale calcChain.xml and mark dynamic-array formulas so
            # Excel 365 spills them without prepending the @ operator.
            _postprocess_xlsx(tmp_path)
            shutil.move(tmp_path, saved_path)
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)
    finally:
        owb.close()

    return saved_path


# ── openpyxl dispatcher ───────────────────────────────────────────────────────

_OPENPYXL_HANDLERS: dict = {}


def _dispatch_openpyxl(owb: Any, change: Change) -> None:
    op = change.type
    handler = _OPENPYXL_HANDLERS.get(op)
    if handler is None:
        raise ApplyError(f"Unknown change type: '{op}'")
    _log.debug("dispatch: %s  params=%s", op, change.params)
    handler(owb, change.params)


# ── compact-flag helpers (AI write-back syntax) ──────────────────────────────

_FLAG_BOOL: dict[str, str] = {
    "B": "bold", "I": "italic", "S": "strikethrough",
    "U": "underline", "W": "wrap_text",
}
_FLAG_H_ALIGN: dict[str, str] = {
    "l": "left", "c": "center", "r": "right", "f": "fill", "j": "justify",
}
_FLAG_V_ALIGN: dict[str, str] = {"t": "top", "c": "center", "b": "bottom"}

_ALL_FMT_KEYS = frozenset({
    "bold", "italic", "strikethrough", "underline", "wrap_text",
    "number_format", "font_name", "font_size", "font_color", "bg_color",
    "h_align", "v_align",
    "border_top", "border_bottom", "border_left", "border_right",
})

_COLOR_KEYS = frozenset({"font_color", "bg_color"})


def _to_argb(color: str) -> str:
    """Ensure a color string is 8-char ARGB (openpyxl requirement).

    Accepts 6-char RGB ('FF0000') or 8-char ARGB ('FFFF0000').
    Empty string is preserved -- callers use it to mean "clear the colour".
    Raises ValueError for anything else so the problem surfaces immediately.
    """
    c = (color or "").strip().upper().lstrip("#")
    if c == "":
        return ""
    if len(c) == 6:
        return f"FF{c}"
    if len(c) == 8:
        return c
    raise ValueError(f"Invalid color value {color!r}: expected 6 or 8 hex chars")


def _normalize_colors(p: dict) -> dict:
    """Return a copy of p with all color keys normalised to 8-char ARGB.

    An empty-string colour is preserved (it signals "clear the colour").
    """
    out = dict(p)
    for key in _COLOR_KEYS:
        if key in out:
            out[key] = _to_argb(out[key])
    return out


def _expand_flags(p: dict) -> dict:
    """Parse compact 'flags' string into format kwargs and merge with p.

    The AI may send: {"type":"set_cell",...,"flags":"B,#$#,##0,bt:thin,~D9E1F2"}
    which is equivalent to explicit keys bold=True, number_format="$#,##0", etc.
    Colors are normalised to 8-char ARGB at parse time.
    """
    flags_str = p.get("flags")
    if not flags_str:
        return _normalize_colors(p)
    extra: dict = {}
    for flag in str(flags_str).split(","):
        flag = flag.strip()
        if not flag:
            continue
        if flag in _FLAG_BOOL:
            extra[_FLAG_BOOL[flag]] = True
        elif flag.startswith("#"):
            extra["number_format"] = flag[1:]
        elif flag.startswith("^"):
            extra["font_color"] = _to_argb(flag[1:])
        elif flag.startswith("~"):
            extra["bg_color"] = _to_argb(flag[1:])
        elif flag.startswith("fn:"):
            extra["font_name"] = flag[3:]
        elif flag.startswith("fs:"):
            try:
                extra["font_size"] = float(flag[3:])
            except ValueError:
                pass
        elif flag.startswith("ha:"):
            extra["h_align"] = _FLAG_H_ALIGN.get(flag[3:], flag[3:])
        elif flag.startswith("va:"):
            extra["v_align"] = _FLAG_V_ALIGN.get(flag[3:], flag[3:])
        elif flag.startswith("bt:"):
            extra["border_top"] = flag[3:]
        elif flag.startswith("bb:"):
            extra["border_bottom"] = flag[3:]
        elif flag.startswith("bl:"):
            extra["border_left"] = flag[3:]
        elif flag.startswith("br:"):
            extra["border_right"] = flag[3:]
    result = {**_normalize_colors(p), **extra}
    result.pop("flags", None)
    return result


# ── cell helpers ──────────────────────────────────────────────────────────────

def _get_openpyxl_sheet(owb: Any, name: str) -> Any:
    try:
        return owb[name]
    except KeyError as exc:
        raise ApplyError(f"Sheet not found: '{name}'") from exc


def _set_cell(owb: Any, p: dict) -> None:
    p = _expand_flags(p)
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    cell = ws[p["cell"]]
    if "formula" in p:
        cell.value = _add_xlfn_prefix(p["formula"])
    else:
        cell.value = p["value"]
    # Apply any inline format keys
    if _ALL_FMT_KEYS & p.keys():
        _set_format(owb, {"sheet": p["sheet"], "range": p["cell"],
                          **{k: p[k] for k in _ALL_FMT_KEYS if k in p}})


def _set_range(owb: Any, p: dict) -> None:
    p = _expand_flags(p)
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    import openpyxl.utils
    rng = p["range"]
    min_col_letter, min_row, max_col_letter, max_row = openpyxl.utils.range_boundaries(rng)
    assert min_col_letter is not None and min_row is not None
    for r_idx, row_data in enumerate(p["values"]):
        for c_idx, val in enumerate(row_data):
            ws.cell(row=min_row + r_idx, column=min_col_letter + c_idx, value=val)
    # Apply any inline format to the entire range
    if _ALL_FMT_KEYS & p.keys():
        _set_format(owb, {"sheet": p["sheet"], "range": rng,
                          **{k: p[k] for k in _ALL_FMT_KEYS if k in p}})


def _clear_range(owb: Any, p: dict) -> None:
    import openpyxl.utils
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(p["range"])
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None


# ── sheet structure ───────────────────────────────────────────────────────────

def _add_sheet(owb: Any, p: dict) -> None:
    name: str = p["name"]
    position: int = p.get("position", len(owb.sheetnames) + 1)
    position = max(1, min(position, len(owb.sheetnames) + 1))
    owb.create_sheet(title=name, index=position - 1)


def _delete_sheet(owb: Any, p: dict) -> None:
    ws = _get_openpyxl_sheet(owb, p["name"])
    del owb[ws.title]


def _rename_sheet(owb: Any, p: dict) -> None:
    ws = _get_openpyxl_sheet(owb, p["old_name"])
    ws.title = p["new_name"]


def _move_sheet(owb: Any, p: dict) -> None:
    ws = _get_openpyxl_sheet(owb, p["name"])
    target = max(0, p["position"] - 1)
    owb.move_sheet(ws, offset=target - owb.index(ws))


def _copy_sheet(owb: Any, p: dict) -> None:
    source = _get_openpyxl_sheet(owb, p["source"])
    idx = p.get("position", len(owb.sheetnames) + 1) - 1
    idx = max(0, min(idx, len(owb.sheetnames)))
    new_ws = owb.copy_worksheet(source)
    new_ws.title = p["dest"]
    owb.move_sheet(new_ws, offset=idx - owb.index(new_ws))


def _hide_sheet(owb: Any, p: dict) -> None:
    ws = _get_openpyxl_sheet(owb, p["name"])
    ws.sheet_state = "hidden"


def _unhide_sheet(owb: Any, p: dict) -> None:
    ws = _get_openpyxl_sheet(owb, p["name"])
    ws.sheet_state = "visible"


# ── table operations ─────────────────────────────────────────────────────────

def _create_table(owb: Any, p: dict) -> None:
    """Create an Excel Table (ListObject) on a worksheet.

    params: sheet, range, name (table display name), style (optional,
    default 'TableStyleMedium9').
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    table_name = p["name"]
    # Remove existing table with same name if present (idempotent)
    if table_name in {t.displayName for t in ws.tables.values()}:
        del ws.tables[table_name]
    style_name = p.get("style", "TableStyleMedium9")
    tbl = Table(displayName=table_name, ref=p["range"])
    tbl.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)


def _delete_table(owb: Any, p: dict) -> None:
    """Remove an Excel Table by name from its sheet."""
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    table_name = p["name"]
    if table_name not in ws.tables:
        raise ApplyError(f"Table '{table_name}' not found on sheet '{p['sheet']}'.")
    del ws.tables[table_name]


# ── named range operations ────────────────────────────────────────────────────

def _add_named_range(owb: Any, p: dict) -> None:
    from openpyxl.workbook.defined_name import DefinedName
    defn = DefinedName(name=p["name"], attr_text=p["refers_to"])
    owb.defined_names[p["name"]] = defn


def _delete_named_range(owb: Any, p: dict) -> None:
    name = p["name"]
    if name not in owb.defined_names:
        raise ApplyError(f"Named range '{name}' not found.")
    del owb.defined_names[name]


# ── merge operations ──────────────────────────────────────────────────────────

def _merge_cells(owb: Any, p: dict) -> None:
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    new_range = p["range"]

    # Overlapping merged ranges produce invalid OOXML and Excel will show a
    # "Removed Records: Merged Cells" repair dialog on open. If the requested
    # range overlaps any existing merge, unmerge those first so the new merge
    # cleanly replaces them.
    from openpyxl.utils import range_boundaries
    new_bounds = range_boundaries(new_range)
    if new_bounds is None:
        ws.merge_cells(new_range)
        return
    n_min_col, n_min_row, n_max_col, n_max_row = new_bounds

    overlapping: list[str] = []
    for existing in list(ws.merged_cells.ranges):
        e_str = str(existing)
        if e_str == new_range:
            return  # identical merge already exists; nothing to do
        e_bounds = range_boundaries(e_str)
        if e_bounds is None:
            continue
        e_min_col, e_min_row, e_max_col, e_max_row = e_bounds
        if (
            n_min_col <= e_max_col and n_max_col >= e_min_col
            and n_min_row <= e_max_row and n_max_row >= e_min_row
        ):
            overlapping.append(e_str)

    for r in overlapping:
        ws.unmerge_cells(r)

    ws.merge_cells(new_range)


def _unmerge_cells(owb: Any, p: dict) -> None:
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    ws.unmerge_cells(p["range"])


# ── format operation ──────────────────────────────────────────────────────────

def _set_format(owb: Any, p: dict) -> None:
    """Apply formatting to a cell or range.

    params: sheet, range, and any subset of: number_format, bold, italic,
    underline, font_name, font_size, font_color, bg_color, h_align, v_align,
    wrap_text.
    """
    import openpyxl.utils
    from openpyxl.styles import Alignment, Font, PatternFill
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    rng = p["range"]
    # Expand to all cells in range
    try:
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(rng)
        assert min_col is not None and min_row is not None and max_col is not None and max_row is not None
        # Whole-column refs ("A:H") and whole-row refs ("1:3") expand to the
        # Excel sheet limits (1,048,576 rows / 16,384 cols). Iterating those
        # is hopeless, so clamp them down to the actual used area. Explicit
        # finite ranges (e.g. "Q2:S6") must be honoured even when they
        # extend past the currently-stored data -- a dynamic-array spill
        # rectangle, for instance, only stores its anchor cell, and the
        # caller still expects formatting to land on every cell in the
        # rectangle.
        if max_row >= 1_048_576:
            max_row = max(min_row, ws.max_row or min_row)
        if max_col >= 16_384:
            max_col = max(min_col, ws.max_column or min_col)
        cells = [ws.cell(row=r, column=c)
                 for r in range(min_row, max_row + 1)
                 for c in range(min_col, max_col + 1)]
    except Exception:  # noqa: BLE001
        # ws[rng] for multi-cell refs returns tuples of tuples; flatten them
        raw = ws[rng]
        if isinstance(raw, tuple):
            cells = []
            for item in raw:
                if isinstance(item, tuple):
                    cells.extend(item)
                else:
                    cells.append(item)
        else:
            cells = [raw]

    for cell in cells:
        if "number_format" in p:
            cell.number_format = p["number_format"]
        if any(k in p for k in ("bold", "italic", "strikethrough", "underline", "font_name", "font_size", "font_color")):
            existing = cell.font or Font()
            fc = p.get("font_color")  # already ARGB from _expand_flags / _normalize_colors
            # Empty-string font_color means "reset to default" (no color).
            if fc == "":
                resolved_color = None
            elif fc is not None:
                resolved_color = fc
            else:
                resolved_color = existing.color or None
            cell.font = Font(
                bold=p.get("bold", existing.bold),
                italic=p.get("italic", existing.italic),
                strike=p.get("strikethrough", existing.strike),
                underline="single" if p.get("underline", existing.underline) else None,
                name=p.get("font_name", existing.name),
                size=p.get("font_size", existing.size),
                color=resolved_color,
            )
        if "bg_color" in p:
            bg = p["bg_color"]
            if bg:
                cell.fill = PatternFill(fill_type="solid", fgColor=bg)  # already ARGB
            else:
                # Empty string means "remove the fill" (no PatternFill type).
                cell.fill = PatternFill(fill_type=None)
        if any(k in p for k in ("h_align", "v_align", "wrap_text")):
            existing_a = cell.alignment or Alignment()
            cell.alignment = Alignment(
                horizontal=p.get("h_align", existing_a.horizontal),
                vertical=p.get("v_align", existing_a.vertical),
                wrap_text=p.get("wrap_text", existing_a.wrap_text),
            )
        if any(k in p for k in ("border_top", "border_bottom", "border_left", "border_right")):
            from openpyxl.styles import Border, Side

            def _make_side(spec: str, existing_side: Any) -> Any:
                if not spec:
                    return existing_side
                parts = spec.split(":", 1)
                style = parts[0]
                color = parts[1] if len(parts) > 1 else "FF000000"
                if len(color) == 6:
                    color = "FF" + color
                return Side(border_style=style, color=color)

            existing_b = cell.border or Border()
            cell.border = Border(
                top=_make_side(p.get("border_top", ""), existing_b.top),
                bottom=_make_side(p.get("border_bottom", ""), existing_b.bottom),
                left=_make_side(p.get("border_left", ""), existing_b.left),
                right=_make_side(p.get("border_right", ""), existing_b.right),
            )


# ── sheet view / dimension operations ────────────────────────────────────────

def _freeze_panes(owb: Any, p: dict) -> None:
    """Freeze rows/columns at a cell.  cell='' or absent to unfreeze."""
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    cell = (p.get("cell") or "").strip() or None
    ws.freeze_panes = cell


def _set_col_width(owb: Any, p: dict) -> None:
    """Set column width.  columns='A' or 'A:D', width in Excel character units."""
    from openpyxl.utils import column_index_from_string, get_column_letter
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    width = float(p["width"])
    col_spec = str(p["columns"]).strip().upper()
    if ":" in col_spec:
        start_col, end_col = col_spec.split(":", 1)
        for idx in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    else:
        ws.column_dimensions[col_spec].width = width


def _set_row_height(owb: Any, p: dict) -> None:
    """Set row height in points."""
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    ws.row_dimensions[int(p["row"])].height = float(p["height"])


def _insert_rows(owb: Any, p: dict) -> None:
    """Insert blank rows above 'row' (1-based)."""
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    ws.insert_rows(int(p["row"]), int(p.get("count", 1)))


def _delete_rows(owb: Any, p: dict) -> None:
    """Delete rows starting at 'row' (1-based)."""
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    ws.delete_rows(int(p["row"]), int(p.get("count", 1)))


def _insert_cols(owb: Any, p: dict) -> None:
    """Insert blank columns before 'col' (letter, e.g. 'B')."""
    from openpyxl.utils import column_index_from_string
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    ws.insert_cols(column_index_from_string(str(p["col"]).strip().upper()), int(p.get("count", 1)))


def _delete_cols(owb: Any, p: dict) -> None:
    """Delete columns starting at 'col' (letter)."""
    from openpyxl.utils import column_index_from_string
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    ws.delete_cols(column_index_from_string(str(p["col"]).strip().upper()), int(p.get("count", 1)))


def _set_tab_color(owb: Any, p: dict) -> None:
    """Set the sheet tab color.  color is RRGGBB or AARRGGBB.  '' to clear."""
    from openpyxl.styles.colors import Color
    ws = _get_openpyxl_sheet(owb, p["name"])
    color_str = (p.get("color") or "").strip()
    ws.sheet_properties.tabColor = Color(rgb=_to_argb(color_str)) if color_str else None


def _set_auto_filter(owb: Any, p: dict) -> None:
    """Set or clear the auto-filter on a sheet.  range='' to clear."""
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    rng = (p.get("range") or "").strip()
    ws.auto_filter.ref = rng or None


def _hide_rows(owb: Any, p: dict) -> None:
    """Hide rows.  row = 1-based start row, count = number of rows (default 1)."""
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    for r in range(int(p["row"]), int(p["row"]) + int(p.get("count", 1))):
        ws.row_dimensions[r].hidden = True


def _unhide_rows(owb: Any, p: dict) -> None:
    """Unhide rows."""
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    for r in range(int(p["row"]), int(p["row"]) + int(p.get("count", 1))):
        ws.row_dimensions[r].hidden = False


def _hide_cols(owb: Any, p: dict) -> None:
    """Hide columns.  col = letter (e.g. 'B'), count = number of cols (default 1)."""
    from openpyxl.utils import column_index_from_string, get_column_letter
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    start = column_index_from_string(str(p["col"]).strip().upper())
    for idx in range(start, start + int(p.get("count", 1))):
        ws.column_dimensions[get_column_letter(idx)].hidden = True


def _unhide_cols(owb: Any, p: dict) -> None:
    """Unhide columns."""
    from openpyxl.utils import column_index_from_string, get_column_letter
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    start = column_index_from_string(str(p["col"]).strip().upper())
    for idx in range(start, start + int(p.get("count", 1))):
        ws.column_dimensions[get_column_letter(idx)].hidden = False


def _clear_format(owb: Any, p: dict) -> None:
    """Reset all formatting on a cell range to defaults (values are untouched)."""
    import openpyxl.utils
    from openpyxl.styles import Alignment, Border, Font, PatternFill
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    try:
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(p["range"])
        assert min_col is not None and min_row is not None and max_col is not None and max_row is not None
        cells = [ws.cell(row=r, column=c)
                 for r in range(min_row, max_row + 1)
                 for c in range(min_col, max_col + 1)]
    except Exception:  # noqa: BLE001
        cells = [ws[p["range"]]]
    for cell in cells:
        cell.font = Font()
        cell.fill = PatternFill()
        cell.border = Border()
        cell.alignment = Alignment()
        cell.number_format = "General"


def _protect_sheet(owb: Any, p: dict) -> None:
    """Password-protect a sheet.  password='' for no password."""
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    password = (p.get("password") or "").strip()
    ws.protection.sheet = True
    if password:
        ws.protection.set_password(password)


def _unprotect_sheet(owb: Any, p: dict) -> None:
    """Remove sheet protection."""
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    ws.protection.sheet = False


def _set_print_area(owb: Any, p: dict) -> None:
    """Set the print area for a sheet.  range='' to clear."""
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    rng = (p.get("range") or "").strip()
    ws.print_area = rng or None


def _set_zoom(owb: Any, p: dict) -> None:
    """Set the sheet zoom level (10-400)."""
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    ws.sheet_view.zoomScale = max(10, min(400, int(p["zoom"])))


# ── chart operations ──────────────────────────────────────────────────────────

_VALID_CHART_TYPES = frozenset(
    {"bar", "col", "column", "line", "pie", "doughnut", "scatter", "area", "radar"}
)


def _create_chart(owb: Any, p: dict) -> None:
    """Create a chart on a worksheet.

    params:
      sheet       - sheet name
      chart_type  - bar | col | line | pie | doughnut | scatter | area | radar
      data_range  - e.g. "A1:D10" or "Sheet2!A1:D10"
                    First column = categories/x-axis, first row = series labels.
      anchor      - top-left cell for chart placement, e.g. "F2"
      title       - chart title (optional)
      x_axis_title, y_axis_title - axis labels (optional, not used for pie/doughnut)
      grouping    - bar/col only: clustered | stacked | percentStacked (default clustered)
      width       - width in cm (default 15)
      height      - height in cm (default 10)
    """
    import openpyxl.utils
    from openpyxl.chart import Reference

    ws = _get_openpyxl_sheet(owb, p["sheet"])
    chart_type = str(p["chart_type"]).lower().strip()
    if chart_type == "column":
        chart_type = "col"
    if chart_type not in _VALID_CHART_TYPES:
        raise ApplyError(
            f"Unknown chart_type '{chart_type}'. "
            "Valid types: bar, col, line, pie, doughnut, scatter, area, radar"
        )

    # Build the chart object
    if chart_type in ("bar", "col"):
        from openpyxl.chart import BarChart
        chart: Any = BarChart()
        chart.type = chart_type
        chart.grouping = p.get("grouping", "clustered")
    elif chart_type == "line":
        from openpyxl.chart import LineChart
        chart = LineChart()
    elif chart_type == "pie":
        from openpyxl.chart import PieChart
        chart = PieChart()
    elif chart_type == "doughnut":
        from openpyxl.chart import DoughnutChart
        chart = DoughnutChart()
    elif chart_type == "scatter":
        from openpyxl.chart import ScatterChart
        chart = ScatterChart()
        chart.style = 10
    elif chart_type == "area":
        from openpyxl.chart import AreaChart
        chart = AreaChart()
    else:  # radar
        from openpyxl.chart import RadarChart
        chart = RadarChart()
        chart.type = "standard"

    if p.get("title"):
        chart.title = str(p["title"])
    chart.width = float(p.get("width", 15))
    chart.height = float(p.get("height", 10))

    if chart_type not in ("pie", "doughnut"):
        if p.get("x_axis_title") and hasattr(chart, "x_axis"):
            chart.x_axis.title = str(p["x_axis_title"])
        if p.get("y_axis_title") and hasattr(chart, "y_axis"):
            chart.y_axis.title = str(p["y_axis_title"])

    # Resolve the data range (optional sheet prefix)
    data_range = str(p["data_range"])
    ref_ws = ws
    if "!" in data_range:
        ref_sheet_name, data_range = data_range.split("!", 1)
        ref_ws = _get_openpyxl_sheet(owb, ref_sheet_name.strip("'\""))

    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(data_range)
    if min_col is None or min_row is None or max_col is None or max_row is None:
        raise ApplyError(f"Invalid data_range '{p['data_range']}'.")
    if max_col <= min_col:
        raise ApplyError(
            f"data_range '{p['data_range']}' must span at least 2 columns "
            "(first column = categories, remaining columns = data series)."
        )

    if chart_type == "scatter":
        data_ref = Reference(
            ref_ws, min_col=min_col + 1, min_row=min_row, max_col=max_col, max_row=max_row
        )
        chart.add_data(data_ref, titles_from_data=True)
        cats_ref = Reference(ref_ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
        chart.set_categories(cats_ref)
    else:
        # data = columns 2..N with header row; categories = column 1 data rows
        data_ref = Reference(
            ref_ws, min_col=min_col + 1, min_row=min_row, max_col=max_col, max_row=max_row
        )
        chart.add_data(data_ref, titles_from_data=True)
        cats_ref = Reference(ref_ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
        chart.set_categories(cats_ref)

    ws.add_chart(chart, str(p.get("anchor", "A1")))


def _delete_chart(owb: Any, p: dict) -> None:
    """Remove a chart from a sheet, identified by title.

    params: sheet, title
    """
    ws = _get_openpyxl_sheet(owb, p["sheet"])
    title = str(p.get("title", "")).strip()
    charts = list(getattr(ws, "_charts", []))
    if not charts:
        raise ApplyError(f"No charts found on sheet '{p['sheet']}'.")
    kept = []
    removed = 0
    for chart in charts:
        ct = getattr(chart, "title", None)
        chart_title = ct if isinstance(ct, str) else ""
        if not chart_title and ct is not None:
            try:
                chart_title = str(ct.tx.rich.p[0].r[0].t)  # type: ignore[union-attr]
            except (AttributeError, IndexError, TypeError):
                pass
        if title and chart_title == title:
            removed += 1
        else:
            kept.append(chart)
    if removed == 0:
        raise ApplyError(f"Chart titled '{title}' not found on sheet '{p['sheet']}'.")
    ws._charts[:] = kept


# ── handler registry (module-level for performance) ───────────────────────────

_OPENPYXL_HANDLERS.update({
    "set_cell":          _set_cell,
    "set_range":         _set_range,
    "clear_range":       _clear_range,
    "add_sheet":         _add_sheet,
    "delete_sheet":      _delete_sheet,
    "rename_sheet":      _rename_sheet,
    "move_sheet":        _move_sheet,
    "copy_sheet":        _copy_sheet,
    "hide_sheet":        _hide_sheet,
    "unhide_sheet":      _unhide_sheet,
    "create_table":       _create_table,
    "delete_table":       _delete_table,
    "freeze_panes":       _freeze_panes,
    "set_col_width":      _set_col_width,
    "set_row_height":     _set_row_height,
    "insert_rows":        _insert_rows,
    "delete_rows":        _delete_rows,
    "insert_cols":        _insert_cols,
    "delete_cols":        _delete_cols,
    "set_tab_color":      _set_tab_color,
    "auto_filter":        _set_auto_filter,
    "hide_rows":          _hide_rows,
    "unhide_rows":        _unhide_rows,
    "hide_cols":          _hide_cols,
    "unhide_cols":        _unhide_cols,
    "clear_format":       _clear_format,
    "protect_sheet":      _protect_sheet,
    "unprotect_sheet":    _unprotect_sheet,
    "set_print_area":     _set_print_area,
    "set_zoom":           _set_zoom,
    "set_named_range":   _add_named_range,
    "delete_named_range": _delete_named_range,
    "merge_cells":       _merge_cells,
    "unmerge_cells":     _unmerge_cells,
    "set_format":        _set_format,
    "create_chart":      _create_chart,
    "delete_chart":      _delete_chart,
})
