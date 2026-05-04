"""Regression tests for core.excel_reader."""
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

from core.excel_reader import read_workbook


class TestArrayFormulaReading:
    """openpyxl returns ArrayFormula objects for cells with <f t="array">.

    The reader must coerce them into proper formula strings; otherwise the
    AI context shows ``repr(ArrayFormula(...))`` instead of the real
    formula text and the model can't reason about it.
    """

    def _save(self, tmp_path, ws_setup):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws_setup(ws)
        path = tmp_path / "book.xlsx"
        wb.save(path)
        return str(path)

    def test_dynamic_array_xlookup_extracted_as_formula(self, tmp_path):
        def setup(ws):
            ws["A1"] = "key"
            ws["B1"] = "val"
            ws["A2"], ws["B2"] = "x", 1
            ws["A3"], ws["B3"] = "y", 2
            ws["J2"] = "x"
            ws["K2"] = ArrayFormula("K2:K7", "=XLOOKUP(J2:J7, A2:A3, B2:B3)")
        path = self._save(tmp_path, setup)

        wb_data = read_workbook(path)
        cell = wb_data.sheets[0].cells["K2"]

        assert cell.formula.startswith("=")
        assert "XLOOKUP" in cell.formula
        assert "ArrayFormula" not in cell.formula
        # The literal value should not still be the ArrayFormula object.
        assert cell.value is None or isinstance(cell.value, (str, int, float))

    def test_sequence_array_formula_extracted(self, tmp_path):
        def setup(ws):
            ws["Q2"] = ArrayFormula("Q2:S6", "=SEQUENCE(5,3)")
        path = self._save(tmp_path, setup)

        wb_data = read_workbook(path)
        cell = wb_data.sheets[0].cells["Q2"]
        assert "SEQUENCE" in cell.formula
        assert cell.formula.startswith("=")

    def test_plain_string_formula_still_works(self, tmp_path):
        def setup(ws):
            ws["A1"] = "=SUM(B1:B3)"
        path = self._save(tmp_path, setup)

        wb_data = read_workbook(path)
        cell = wb_data.sheets[0].cells["A1"]
        assert cell.formula == "=SUM(B1:B3)"
        assert cell.value is None
