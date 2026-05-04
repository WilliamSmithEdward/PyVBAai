# Copyright (c) 2026 William E. Smith (williamsmithe@icloud.com). All rights reserved.
# Proprietary and confidential. Unauthorized use, copying, or distribution
# of this file, via any medium, is strictly prohibited. See LICENSE for details.
"""Tests for core/excel_writer.py.

All handlers use real openpyxl workbooks created in-memory.
"""
from __future__ import annotations

import openpyxl
import pytest

from core.excel_writer import (
    _DA_FEATURES_URI,
    ApplyError,
    _add_named_range,
    _add_sheet,
    _add_xlfn_prefix,
    _clear_range,
    _copy_sheet,
    _create_chart,
    _delete_chart,
    _delete_named_range,
    _delete_sheet,
    _dispatch_openpyxl,
    _expand_flags,
    _get_openpyxl_sheet,
    _merge_cells,
    _move_sheet,
    _rename_sheet,
    _set_cell,
    _set_format,
    _set_range,
    _to_argb,
    _unmerge_cells,
)
from models.conversation import Change

# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture()
def owb():
    """A real openpyxl Workbook with two sheets for testing."""
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.create_sheet("Sheet2")
    return wb



# ── _get_openpyxl_sheet ───────────────────────────────────────────────────────

class TestGetOpenpyxlSheet:
    def test_returns_sheet(self, owb):
        ws = _get_openpyxl_sheet(owb, "Sheet1")
        assert ws.title == "Sheet1"

    def test_raises_apply_error_on_missing(self, owb):
        with pytest.raises(ApplyError, match="Sheet not found"):
            _get_openpyxl_sheet(owb, "Ghost")


# ── _set_cell ─────────────────────────────────────────────────────────────────

class TestSetCell:
    def test_sets_value(self, owb):
        _set_cell(owb, {"sheet": "Sheet1", "cell": "A1", "value": 42})
        assert owb["Sheet1"]["A1"].value == 42

    def test_sets_formula(self, owb):
        _set_cell(owb, {"sheet": "Sheet1", "cell": "B2", "formula": "=SUM(A:A)"})
        assert owb["Sheet1"]["B2"].value == "=SUM(A:A)"

    def test_formula_takes_priority(self, owb):
        _set_cell(owb, {"sheet": "Sheet1", "cell": "A1", "formula": "=1+1", "value": 99})
        assert owb["Sheet1"]["A1"].value == "=1+1"

    def test_raises_for_unknown_sheet(self, owb):
        with pytest.raises(ApplyError):
            _set_cell(owb, {"sheet": "Ghost", "cell": "A1", "value": 1})


# ── _set_range ────────────────────────────────────────────────────────────────

class TestSetRange:
    def test_sets_values(self, owb):
        _set_range(owb, {"sheet": "Sheet1", "range": "A1:C2",
                         "values": [[1, 2, 3], [4, 5, 6]]})
        ws = owb["Sheet1"]
        assert ws["A1"].value == 1
        assert ws["B2"].value == 5
        assert ws["C2"].value == 6

    def test_single_row(self, owb):
        _set_range(owb, {"sheet": "Sheet1", "range": "A1:C1",
                         "values": [["a", "b", "c"]]})
        ws = owb["Sheet1"]
        assert ws["A1"].value == "a"
        assert ws["C1"].value == "c"


# ── _clear_range ──────────────────────────────────────────────────────────────

class TestClearRange:
    def test_clears_values(self, owb):
        ws = owb["Sheet1"]
        ws["A1"].value = 100
        ws["B2"].value = "hello"
        _clear_range(owb, {"sheet": "Sheet1", "range": "A1:B2"})
        assert ws["A1"].value is None
        assert ws["B2"].value is None


# ── _add_sheet ────────────────────────────────────────────────────────────────

class TestAddSheet:
    def test_add_new_sheet(self, owb):
        _add_sheet(owb, {"name": "NewSheet"})
        assert "NewSheet" in owb.sheetnames

    def test_add_at_position_1(self, owb):
        _add_sheet(owb, {"name": "First", "position": 1})
        assert owb.sheetnames[0] == "First"

    def test_default_appends(self, owb):
        count_before = len(owb.sheetnames)
        _add_sheet(owb, {"name": "Last"})
        assert owb.sheetnames[-1] == "Last"
        assert len(owb.sheetnames) == count_before + 1


# ── _delete_sheet ─────────────────────────────────────────────────────────────

class TestDeleteSheet:
    def test_deletes_sheet(self, owb):
        _delete_sheet(owb, {"name": "Sheet2"})
        assert "Sheet2" not in owb.sheetnames

    def test_raises_for_missing_sheet(self, owb):
        with pytest.raises(ApplyError):
            _delete_sheet(owb, {"name": "Ghost"})


# ── _rename_sheet ─────────────────────────────────────────────────────────────

class TestRenameSheet:
    def test_renames(self, owb):
        _rename_sheet(owb, {"old_name": "Sheet1", "new_name": "Renamed"})
        assert "Renamed" in owb.sheetnames
        assert "Sheet1" not in owb.sheetnames


# ── _move_sheet ───────────────────────────────────────────────────────────────

class TestMoveSheet:
    def test_move_to_position_1(self, owb):
        # owb starts: [Sheet1, Sheet2]
        _move_sheet(owb, {"name": "Sheet2", "position": 1})
        assert owb.sheetnames[0] == "Sheet2"

    def test_move_to_end(self, owb):
        owb.create_sheet("Sheet3")
        _move_sheet(owb, {"name": "Sheet1", "position": 3})
        assert owb.sheetnames[-1] == "Sheet1"


# ── _copy_sheet ───────────────────────────────────────────────────────────────

class TestCopySheet:
    def test_copies_sheet(self, owb):
        owb["Sheet1"]["A1"].value = "orig"
        _copy_sheet(owb, {"source": "Sheet1", "dest": "CopyOfSheet1", "position": 3})
        assert "CopyOfSheet1" in owb.sheetnames

    def test_raises_for_missing_source(self, owb):
        with pytest.raises(ApplyError):
            _copy_sheet(owb, {"source": "Ghost", "dest": "Copy"})


# ── _merge_cells / _unmerge_cells ─────────────────────────────────────────────

class TestMergeCells:
    def test_merge(self, owb):
        _merge_cells(owb, {"sheet": "Sheet1", "range": "A1:B2"})
        assert "A1:B2" in [str(m) for m in owb["Sheet1"].merged_cells.ranges]

    def test_unmerge(self, owb):
        owb["Sheet1"].merge_cells("C1:D2")
        _unmerge_cells(owb, {"sheet": "Sheet1", "range": "C1:D2"})
        assert "C1:D2" not in [str(m) for m in owb["Sheet1"].merged_cells.ranges]

    def test_overlapping_merge_replaces_existing(self, owb):
        # Excel rejects overlapping merged ranges with a "Removed Records:
        # Merged Cells" repair dialog, so the second merge must unmerge any
        # overlapping ranges first.
        _merge_cells(owb, {"sheet": "Sheet1", "range": "A1:J1"})
        _merge_cells(owb, {"sheet": "Sheet1", "range": "A1:Y1"})
        ranges = [str(m) for m in owb["Sheet1"].merged_cells.ranges]
        assert "A1:J1" not in ranges
        assert "A1:Y1" in ranges

    def test_overlapping_merge_replaces_multiple(self, owb):
        # A new merge that covers two existing non-adjacent merges should
        # replace both of them.
        _merge_cells(owb, {"sheet": "Sheet1", "range": "A1:B2"})
        _merge_cells(owb, {"sheet": "Sheet1", "range": "D1:E2"})
        _merge_cells(owb, {"sheet": "Sheet1", "range": "A1:F3"})
        ranges = [str(m) for m in owb["Sheet1"].merged_cells.ranges]
        assert "A1:B2" not in ranges
        assert "D1:E2" not in ranges
        assert "A1:F3" in ranges

    def test_non_overlapping_merge_preserved(self, owb):
        _merge_cells(owb, {"sheet": "Sheet1", "range": "A1:B2"})
        _merge_cells(owb, {"sheet": "Sheet1", "range": "D4:E5"})
        ranges = [str(m) for m in owb["Sheet1"].merged_cells.ranges]
        assert "A1:B2" in ranges
        assert "D4:E5" in ranges

    def test_identical_merge_is_noop(self, owb):
        _merge_cells(owb, {"sheet": "Sheet1", "range": "A1:B2"})
        _merge_cells(owb, {"sheet": "Sheet1", "range": "A1:B2"})
        ranges = [str(m) for m in owb["Sheet1"].merged_cells.ranges]
        assert ranges.count("A1:B2") == 1


# ── _add_named_range / _delete_named_range ────────────────────────────────────

class TestNamedRange:
    def test_add_named_range(self, owb):
        _add_named_range(owb, {"name": "MyRange", "refers_to": "Sheet1!$A$1:$D$10"})
        assert "MyRange" in owb.defined_names

    def test_delete_named_range(self, owb):
        from openpyxl.workbook.defined_name import DefinedName
        owb.defined_names["ToDelete"] = DefinedName("ToDelete", attr_text="Sheet1!$A$1")
        _delete_named_range(owb, {"name": "ToDelete"})
        assert "ToDelete" not in owb.defined_names

    def test_delete_missing_raises(self, owb):
        with pytest.raises(ApplyError, match="Named range 'Ghost' not found"):
            _delete_named_range(owb, {"name": "Ghost"})


# ── _dispatch_openpyxl ────────────────────────────────────────────────────────

class TestDispatchOpenpyxl:
    def test_unknown_type_raises(self, owb):
        change = Change(type="do_magic", params={})
        with pytest.raises(ApplyError, match="Unknown change type: 'do_magic'"):
            _dispatch_openpyxl(owb, change)

    @pytest.mark.parametrize("op_type,params", [
        ("set_cell",          {"sheet": "Sheet1", "cell": "A1", "value": 1}),
        ("set_range",         {"sheet": "Sheet1", "range": "A1:B2", "values": [[1, 2], [3, 4]]}),
        ("clear_range",       {"sheet": "Sheet1", "range": "A1:Z100"}),
        ("delete_sheet",      {"name": "Sheet2"}),
        ("rename_sheet",      {"old_name": "Sheet1", "new_name": "Renamed"}),
    ])
    def test_known_types_dispatched(self, owb, op_type, params):
        """Known types should dispatch (may mutate owb) without 'Unknown change type'."""
        change = Change(type=op_type, params=params)
        try:
            _dispatch_openpyxl(owb, change)
        except ApplyError as exc:
            assert "Unknown change type" not in str(exc)


class TestToArgb:
    def test_6_char_prefixed(self):
        assert _to_argb("FF0000") == "FFFF0000"

    def test_8_char_unchanged(self):
        assert _to_argb("FFFF0000") == "FFFF0000"

    def test_lowercase_normalised(self):
        assert _to_argb("ff0000") == "FFFF0000"

    def test_hash_prefix_stripped(self):
        assert _to_argb("#92D050") == "FF92D050"

    def test_invalid_raises(self):
        with pytest.raises(ValueError):
            _to_argb("XYZ")

    def test_too_short_raises(self):
        with pytest.raises(ValueError):
            _to_argb("F00")


class TestExpandFlagsColors:
    def test_tilde_flag_normalised_to_argb(self):
        result = _expand_flags({"flags": "~92D050"})
        assert result["bg_color"] == "FF92D050"

    def test_caret_flag_normalised_to_argb(self):
        result = _expand_flags({"flags": "^FF0000"})
        assert result["font_color"] == "FFFF0000"

    def test_explicit_6char_color_normalised(self):
        result = _expand_flags({"bg_color": "D9E1F2"})
        assert result["bg_color"] == "FFD9E1F2"

    def test_explicit_8char_color_unchanged(self):
        result = _expand_flags({"font_color": "FF123456"})
        assert result["font_color"] == "FF123456"


class TestSetFormatColors:
    def test_bg_color_applied(self, owb):
        ws = owb.active
        _set_format(owb, {"sheet": ws.title, "range": "A1", "bg_color": "92D050"})
        assert ws["A1"].fill.fgColor.rgb[-6:] == "92D050"

    def test_font_color_applied(self, owb):
        ws = owb.active
        _set_format(owb, {"sheet": ws.title, "range": "A1", "font_color": "FF0000"})
        assert ws["A1"].font.color.rgb[-6:] == "FF0000"

    def test_set_cell_with_color_flag(self, owb):
        ws = owb.active
        _set_cell(owb, {"sheet": ws.title, "cell": "B1", "value": "x", "flags": "~92D050"})
        assert ws["B1"].fill.fgColor.rgb[-6:] == "92D050"

    def test_empty_bg_color_clears_fill(self, owb):
        # The AI typically issues a "clear" (bg_color="") before re-colouring
        # a region.  An empty colour string used to raise "Colors must be aRGB
        # hex values"; it must instead drop the fill back to none.
        ws = owb.active
        _set_format(owb, {"sheet": ws.title, "range": "A1", "bg_color": "92D050"})
        assert ws["A1"].fill.fgColor.rgb[-6:] == "92D050"
        _set_format(owb, {"sheet": ws.title, "range": "A1", "bg_color": ""})
        assert ws["A1"].fill.fill_type in (None, "none")

    def test_empty_font_color_resets_to_default(self, owb):
        ws = owb.active
        _set_format(owb, {"sheet": ws.title, "range": "A1", "font_color": "FF0000"})
        _set_format(owb, {"sheet": ws.title, "range": "A1", "font_color": ""})
        # Default font has no explicit color set.
        assert ws["A1"].font.color is None

    def test_range_extending_past_used_area_is_honored(self, owb):
        # Regression: _set_format used to clamp every range to ws.max_row /
        # ws.max_column to defend against whole-column refs like "A:H".  That
        # silently truncated explicit ranges that extend past the stored data
        # area -- e.g. a dynamic-array spill rectangle Q2:S6 where only Q2
        # holds the formula and R/S are populated by Excel on open.
        ws = owb.active
        # Plant data only in column A so ws.max_column == 1 and ws.max_row == 1.
        ws["A1"] = "anchor"
        # Now ask for a 5x3 rectangle that extends well past the used area.
        _set_format(owb, {"sheet": ws.title, "range": "Q2:S6", "bg_color": "E6E0F8"})
        # Every cell in the requested rectangle must have received the fill.
        for col in ("Q", "R", "S"):
            for row in range(2, 7):
                cell = ws[f"{col}{row}"]
                assert cell.fill.fgColor.rgb[-6:] == "E6E0F8", (
                    f"{col}{row} was not coloured -- range was clamped"
                )

    def test_whole_column_ref_still_clamped(self, owb):
        # Make sure the original whole-column safety net still works: "A:H"
        # must NOT iterate 1,048,576 rows.  We only need to verify it returns
        # quickly and that at least the cells inside the actual used area get
        # the fill applied.
        ws = owb.active
        ws["A1"] = "x"
        ws["H1"] = "y"
        _set_format(owb, {"sheet": ws.title, "range": "A:H", "bg_color": "FFFF00"})
        assert ws["A1"].fill.fgColor.rgb[-6:] == "FFFF00"
        assert ws["H1"].fill.fgColor.rgb[-6:] == "FFFF00"


# ── _create_chart / _delete_chart ─────────────────────────────────────────────

def _make_chart_data_wb():
    """Return a workbook with simple chart source data on Sheet1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Month", "Sales", "Cost"])
    ws.append(["Jan", 100, 80])
    ws.append(["Feb", 120, 90])
    ws.append(["Mar", 140, 95])
    return wb


class TestAddXlfnPrefix:
    def test_xlookup_prefixed(self):
        assert _add_xlfn_prefix('=XLOOKUP(A1,B:B,C:C,"")') == '=_xlfn.XLOOKUP(A1,B:B,C:C,"")'

    def test_filter_prefixed(self):
        assert _add_xlfn_prefix("=FILTER(A2:C8,B2:B8>0)") == "=_xlfn.FILTER(A2:C8,B2:B8>0)"

    def test_unique_sort_sequence_prefixed(self):
        assert _add_xlfn_prefix("=UNIQUE(A1:A10)") == "=_xlfn.UNIQUE(A1:A10)"
        assert _add_xlfn_prefix("=SORT(A1:A10)") == "=_xlfn.SORT(A1:A10)"
        assert _add_xlfn_prefix("=SEQUENCE(10)") == "=_xlfn.SEQUENCE(10)"

    def test_already_prefixed_unchanged(self):
        assert _add_xlfn_prefix("=_xlfn.XLOOKUP(A1,B:B,C:C)") == "=_xlfn.XLOOKUP(A1,B:B,C:C)"

    def test_nested_function(self):
        result = _add_xlfn_prefix('=IFERROR(XLOOKUP(A1,B:B,C:C,""),0)')
        assert result == '=IFERROR(_xlfn.XLOOKUP(A1,B:B,C:C,""),0)'

    def test_normal_function_unchanged(self):
        assert _add_xlfn_prefix("=SUM(A1:A10)") == "=SUM(A1:A10)"
        assert _add_xlfn_prefix("=VLOOKUP(A1,B:C,2,0)") == "=VLOOKUP(A1,B:C,2,0)"

    def test_non_formula_unchanged(self):
        assert _add_xlfn_prefix("hello") == "hello"
        assert _add_xlfn_prefix("") == ""

    def test_case_insensitive(self):
        assert _add_xlfn_prefix("=xlookup(A1,B:B,C:C)") == "=_xlfn.XLOOKUP(A1,B:B,C:C)"

    def test_set_cell_applies_prefix(self, owb):
        _set_cell(owb, {"sheet": "Sheet1", "cell": "A1", "formula": "=XLOOKUP(1,B:B,C:C)"})
        assert owb["Sheet1"]["A1"].value == "=_xlfn.XLOOKUP(1,B:B,C:C)"

    def test_spilled_range_operator_rewritten(self):
        # "O2#" is Excel formula-bar syntax for a spilled range; the OOXML
        # form is _xlfn.ANCHORARRAY(O2). Storing the literal "#" makes Excel
        # discard the formula on next open ("Removed Records: Formula").
        result = _add_xlfn_prefix('=XLOOKUP("Apple",O2#,D2:D7,"Not found")')
        assert result == '=_xlfn.XLOOKUP("Apple",_xlfn.ANCHORARRAY(O2),D2:D7,"Not found")'

    def test_spilled_range_with_sheet_qualifier(self):
        # Excel stores "Sheet1!A1#" with the sheet qualifier inside the
        # ANCHORARRAY call: _xlfn.ANCHORARRAY(Sheet1!A1).
        result = _add_xlfn_prefix("=SUM(Sheet1!A1#)")
        assert result == "=SUM(_xlfn.ANCHORARRAY(Sheet1!A1))"

    def test_spilled_range_with_quoted_sheet_qualifier(self):
        result = _add_xlfn_prefix("=SUM('My Sheet'!A1#)")
        assert result == "=SUM(_xlfn.ANCHORARRAY('My Sheet'!A1))"

    def test_spilled_range_absolute(self):
        result = _add_xlfn_prefix("=SUM($A$1#)")
        assert result == "=SUM(_xlfn.ANCHORARRAY($A$1))"

    def test_error_tokens_not_rewritten(self):
        # #REF!, #NAME?, #VALUE! must not be touched.
        assert _add_xlfn_prefix("=A1+#REF!") == "=A1+#REF!"
        assert _add_xlfn_prefix('=IFERROR(A1,"#N/A")') == '=IFERROR(A1,"#N/A")'


class TestDaFeaturesInjection:
    """_postprocess_xlsx produces a fully dynamic-array-aware xlsx."""

    def _make_zip(self, tmp_path, sheets=None, has_metadata=False):
        import zipfile as zf
        xlsx = tmp_path / "wb.xlsx"
        cts = (
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Override PartName="/xl/workbook.xml"'
            ' ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            "</Types>"
        )
        rels = (
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
            "</Relationships>"
        )
        wb_xml = "<workbook></workbook>"
        with zf.ZipFile(xlsx, "w") as z:
            z.writestr("[Content_Types].xml", cts)
            z.writestr("xl/workbook.xml", wb_xml)
            z.writestr("xl/_rels/workbook.xml.rels", rels)
            for name, content in (sheets or {}).items():
                z.writestr(f"xl/worksheets/{name}", content)
            if has_metadata:
                z.writestr("xl/metadata.xml", "<metadata/>")
        return xlsx

    def _run(self, xlsx):
        from core.excel_writer import _postprocess_xlsx
        _postprocess_xlsx(str(xlsx))
        import zipfile as zf
        with zf.ZipFile(xlsx) as z:
            return {n: z.read(n).decode() for n in z.namelist()}

    def test_features_uri_injected_no_extlst(self, tmp_path):
        xlsx = self._make_zip(tmp_path)
        data = self._run(xlsx)
        assert _DA_FEATURES_URI in data["xl/workbook.xml"]
        assert "xcalcf:feature" in data["xl/workbook.xml"]

    def test_features_uri_injected_existing_extlst(self, tmp_path):
        import zipfile as zf
        xlsx = tmp_path / "wb.xlsx"
        xml = '<workbook><extLst><ext uri="{OTHER}"/></extLst></workbook>'
        with zf.ZipFile(xlsx, "w") as z:
            z.writestr("xl/workbook.xml", xml)
            z.writestr("[Content_Types].xml", "<Types></Types>")
            z.writestr("xl/_rels/workbook.xml.rels", "<Relationships></Relationships>")
        data = self._run(xlsx)
        assert "{OTHER}" in data["xl/workbook.xml"]
        assert _DA_FEATURES_URI in data["xl/workbook.xml"]

    def test_not_duplicated_if_already_present(self, tmp_path):
        import zipfile as zf
        xlsx = tmp_path / "wb.xlsx"
        uri = _DA_FEATURES_URI
        xml = f'<workbook><extLst><ext uri="{uri}"/></extLst></workbook>'
        with zf.ZipFile(xlsx, "w") as z:
            z.writestr("xl/workbook.xml", xml)
            z.writestr("[Content_Types].xml", "<Types></Types>")
            z.writestr("xl/_rels/workbook.xml.rels", "<Relationships></Relationships>")
        data = self._run(xlsx)
        assert data["xl/workbook.xml"].count(uri) == 1

    def test_calc_chain_removed(self, tmp_path):
        import zipfile as zf
        xlsx = tmp_path / "wb.xlsx"
        with zf.ZipFile(xlsx, "w") as z:
            z.writestr("xl/workbook.xml", "<workbook></workbook>")
            z.writestr("[Content_Types].xml", "<Types></Types>")
            z.writestr("xl/_rels/workbook.xml.rels", "<Relationships></Relationships>")
            z.writestr("xl/calcChain.xml", "<calcChain/>")
        data = self._run(xlsx)
        assert "xl/calcChain.xml" not in data

    def test_metadata_xml_added(self, tmp_path):
        xlsx = self._make_zip(tmp_path)
        data = self._run(xlsx)
        assert "xl/metadata.xml" in data
        assert "XLDAPR" in data["xl/metadata.xml"]
        assert "dynamicArrayProperties" in data["xl/metadata.xml"]

    def test_metadata_content_type_added(self, tmp_path):
        xlsx = self._make_zip(tmp_path)
        data = self._run(xlsx)
        assert "sheetMetadata" in data["[Content_Types].xml"]

    def test_metadata_relationship_added(self, tmp_path):
        xlsx = self._make_zip(tmp_path)
        data = self._run(xlsx)
        assert "sheetMetadata" in data["xl/_rels/workbook.xml.rels"]

    def test_metadata_not_duplicated(self, tmp_path):
        xlsx = self._make_zip(tmp_path, has_metadata=True)
        data = self._run(xlsx)
        # Already had metadata.xml -- should not duplicate
        assert data["xl/metadata.xml"] == "<metadata/>"

    def test_da_cell_cm_and_t_array_added(self, tmp_path):
        sheet = (
            '<worksheet><sheetData>'
            '<row r="1"><c r="A1"><f>_xlfn.UNIQUE(B2:B8)</f><v/></c></row>'
            '</sheetData></worksheet>'
        )
        xlsx = self._make_zip(tmp_path, sheets={"sheet1.xml": sheet})
        data = self._run(xlsx)
        ws = data["xl/worksheets/sheet1.xml"]
        assert 'cm="1"' in ws
        assert 't="array"' in ws
        assert 'ref="A1"' in ws

    def test_plain_formula_not_modified(self, tmp_path):
        sheet = (
            '<worksheet><sheetData>'
            '<row r="1"><c r="A1"><f>SUM(B2:B8)</f><v/></c></row>'
            '</sheetData></worksheet>'
        )
        xlsx = self._make_zip(tmp_path, sheets={"sheet1.xml": sheet})
        data = self._run(xlsx)
        ws = data["xl/worksheets/sheet1.xml"]
        assert 'cm="1"' not in ws
        assert 't="array"' not in ws

    def test_already_marked_not_duplicated(self, tmp_path):
        sheet = (
            '<worksheet><sheetData>'
            '<row r="1"><c r="A1" cm="1"><f t="array" ref="A1">_xlfn.UNIQUE(B2:B8)</f><v/></c></row>'
            '</sheetData></worksheet>'
        )
        xlsx = self._make_zip(tmp_path, sheets={"sheet1.xml": sheet})
        data = self._run(xlsx)
        ws = data["xl/worksheets/sheet1.xml"]
        assert ws.count('cm="1"') == 1
        assert ws.count('t="array"') == 1

    def test_existing_da_array_formula_gets_cm(self, tmp_path):
        # Simulates what openpyxl emits when re-saving a workbook that
        # already had a dynamic-array formula: <f t="array" ref="..."> is
        # preserved but the cm="1" cell-metadata reference is dropped.
        # The postprocessor must restore cm="1" so Excel keeps it as a
        # dynamic array instead of converting it to a legacy CSE array.
        sheet = (
            '<worksheet><sheetData>'
            '<row r="1"><c r="C1"><f t="array" ref="C1">_xlfn.UNIQUE(A1:A3)</f><v/></c></row>'
            '</sheetData></worksheet>'
        )
        xlsx = self._make_zip(tmp_path, sheets={"sheet1.xml": sheet})
        data = self._run(xlsx)
        ws = data["xl/worksheets/sheet1.xml"]
        assert 'cm="1"' in ws
        # original t="array" ref="C1" must be preserved exactly once
        assert ws.count('t="array"') == 1
        assert 'ref="C1"' in ws

    def test_existing_da_with_spill_range_preserved(self, tmp_path):
        # Re-saved workbook with a multi-cell ref (e.g. C1:C3) must keep
        # that ref untouched while still gaining cm="1".
        sheet = (
            '<worksheet><sheetData>'
            '<row r="1"><c r="C1"><f t="array" ref="C1:C3">_xlfn.SORT(A1:A3)</f><v/></c></row>'
            '</sheetData></worksheet>'
        )
        xlsx = self._make_zip(tmp_path, sheets={"sheet1.xml": sheet})
        data = self._run(xlsx)
        ws = data["xl/worksheets/sheet1.xml"]
        assert 'cm="1"' in ws
        assert 'ref="C1:C3"' in ws

    def test_existing_legacy_cse_array_not_modified(self, tmp_path):
        # A legacy (non-dynamic) array formula like {=SUM(A1:A3*B1:B3)} must
        # NOT be promoted to a dynamic array -- only formulas containing one
        # of the recognised _xlfn dynamic-array functions get cm="1".
        sheet = (
            '<worksheet><sheetData>'
            '<row r="1"><c r="C1"><f t="array" ref="C1">SUM(A1:A3*B1:B3)</f><v/></c></row>'
            '</sheetData></worksheet>'
        )
        xlsx = self._make_zip(tmp_path, sheets={"sheet1.xml": sheet})
        data = self._run(xlsx)
        ws = data["xl/worksheets/sheet1.xml"]
        assert 'cm="1"' not in ws

    def test_openpyxl_roundtrip_preserves_dynamic_array(self, tmp_path):
        # End-to-end: build a workbook through our pipeline, then load and
        # re-save it via plain openpyxl (simulating apply_changes with no
        # changes touching the DA cell), and verify the resulting file still
        # has cm="1" + t="array" + the metadata/features parts after a second
        # _postprocess_xlsx pass.
        import zipfile as zf

        import openpyxl

        from core.excel_writer import _postprocess_xlsx

        # Initial save through the pipeline.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "x"
        ws["A2"] = "y"
        ws["A3"] = "x"
        ws["C1"] = "=_xlfn.UNIQUE(A1:A3)"
        first = tmp_path / "first.xlsx"
        wb.save(str(first))
        _postprocess_xlsx(str(first))

        # Round-trip through openpyxl with NO changes, then postprocess again.
        wb2 = openpyxl.load_workbook(str(first))
        second = tmp_path / "second.xlsx"
        wb2.save(str(second))
        _postprocess_xlsx(str(second))

        with zf.ZipFile(second) as z:
            ws_xml = z.read("xl/worksheets/sheet1.xml").decode()
            assert "xl/metadata.xml" in z.namelist()
            wb_xml = z.read("xl/workbook.xml").decode()

        assert 'cm="1"' in ws_xml
        assert 't="array"' in ws_xml
        assert _DA_FEATURES_URI in wb_xml


class TestCreateChart:
    def test_column_chart_added(self):
        wb = _make_chart_data_wb()
        _create_chart(wb, {
            "sheet": "Sheet1",
            "chart_type": "col",
            "data_range": "A1:C4",
            "anchor": "E2",
            "title": "Sales Chart",
        })
        assert len(wb["Sheet1"]._charts) == 1

    def test_bar_chart_type_attribute(self):
        wb = _make_chart_data_wb()
        _create_chart(wb, {
            "sheet": "Sheet1",
            "chart_type": "bar",
            "data_range": "A1:C4",
            "anchor": "E2",
        })
        chart = wb["Sheet1"]._charts[0]
        assert chart.type == "bar"

    def test_line_chart_added(self):
        wb = _make_chart_data_wb()
        _create_chart(wb, {
            "sheet": "Sheet1",
            "chart_type": "line",
            "data_range": "A1:C4",
            "anchor": "E2",
        })
        from openpyxl.chart import LineChart
        assert isinstance(wb["Sheet1"]._charts[0], LineChart)

    def test_pie_chart_added(self):
        wb = _make_chart_data_wb()
        # pie uses only 2 columns: labels + values
        _create_chart(wb, {
            "sheet": "Sheet1",
            "chart_type": "pie",
            "data_range": "A1:B4",
            "anchor": "E2",
        })
        from openpyxl.chart import PieChart
        assert isinstance(wb["Sheet1"]._charts[0], PieChart)

    def test_scatter_chart_added(self):
        wb = _make_chart_data_wb()
        _create_chart(wb, {
            "sheet": "Sheet1",
            "chart_type": "scatter",
            "data_range": "A1:C4",
            "anchor": "E2",
        })
        from openpyxl.chart import ScatterChart
        assert isinstance(wb["Sheet1"]._charts[0], ScatterChart)

    def test_title_set(self):
        wb = _make_chart_data_wb()
        _create_chart(wb, {
            "sheet": "Sheet1",
            "chart_type": "col",
            "data_range": "A1:C4",
            "anchor": "E2",
            "title": "My Chart",
        })
        # openpyxl wraps the title string in a Title object; verify it is set
        chart = wb["Sheet1"]._charts[0]
        t = chart.title
        title_text = str(t) if isinstance(t, str) else t.tx.rich.p[0].r[0].t
        assert title_text == "My Chart"

    def test_size_set(self):
        wb = _make_chart_data_wb()
        _create_chart(wb, {
            "sheet": "Sheet1",
            "chart_type": "line",
            "data_range": "A1:C4",
            "anchor": "E2",
            "width": 20,
            "height": 12,
        })
        chart = wb["Sheet1"]._charts[0]
        assert chart.width == 20
        assert chart.height == 12

    def test_unknown_type_raises(self):
        wb = _make_chart_data_wb()
        with pytest.raises(ApplyError, match="Unknown chart_type"):
            _create_chart(wb, {
                "sheet": "Sheet1",
                "chart_type": "donut_invalid",
                "data_range": "A1:C4",
                "anchor": "E2",
            })

    def test_single_column_range_raises(self):
        wb = _make_chart_data_wb()
        with pytest.raises(ApplyError, match="at least 2 columns"):
            _create_chart(wb, {
                "sheet": "Sheet1",
                "chart_type": "col",
                "data_range": "A1:A4",
                "anchor": "E2",
            })

    def test_unknown_sheet_raises(self):
        wb = _make_chart_data_wb()
        with pytest.raises(ApplyError, match="Sheet not found"):
            _create_chart(wb, {
                "sheet": "Ghost",
                "chart_type": "col",
                "data_range": "A1:C4",
                "anchor": "E2",
            })

    def test_column_alias_works(self):
        wb = _make_chart_data_wb()
        _create_chart(wb, {
            "sheet": "Sheet1",
            "chart_type": "column",
            "data_range": "A1:C4",
            "anchor": "E2",
        })
        from openpyxl.chart import BarChart
        chart = wb["Sheet1"]._charts[0]
        assert isinstance(chart, BarChart)
        assert chart.type == "col"

    def test_stacked_grouping(self):
        wb = _make_chart_data_wb()
        _create_chart(wb, {
            "sheet": "Sheet1",
            "chart_type": "col",
            "data_range": "A1:C4",
            "anchor": "E2",
            "grouping": "stacked",
        })
        assert wb["Sheet1"]._charts[0].grouping == "stacked"


class TestDeleteChart:
    def test_delete_by_title(self):
        wb = _make_chart_data_wb()
        _create_chart(wb, {
            "sheet": "Sheet1", "chart_type": "col",
            "data_range": "A1:C4", "anchor": "E2", "title": "Remove Me",
        })
        assert len(wb["Sheet1"]._charts) == 1
        _delete_chart(wb, {"sheet": "Sheet1", "title": "Remove Me"})
        assert len(wb["Sheet1"]._charts) == 0

    def test_delete_wrong_title_raises(self):
        wb = _make_chart_data_wb()
        _create_chart(wb, {
            "sheet": "Sheet1", "chart_type": "col",
            "data_range": "A1:C4", "anchor": "E2", "title": "Keep Me",
        })
        with pytest.raises(ApplyError, match="not found"):
            _delete_chart(wb, {"sheet": "Sheet1", "title": "Wrong Title"})
        assert len(wb["Sheet1"]._charts) == 1

    def test_delete_no_charts_raises(self):
        wb = _make_chart_data_wb()
        with pytest.raises(ApplyError, match="No charts found"):
            _delete_chart(wb, {"sheet": "Sheet1", "title": "Any"})
